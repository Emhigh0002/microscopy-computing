import os
import csv
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from app.core.config import settings

class ReportService:
    def generate_pdf_report(self, file_path: str, data: Dict[str, Any]) -> str:
        """
        Generates a professional clinical PDF report using ReportLab.
        """
        doc = SimpleDocTemplate(file_path, pagesize=letter,
                                rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=15
        )
        
        meta_style = ParagraphStyle(
            'ReportMeta',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#4b5563'),
            spaceAfter=6
        )

        section_style = ParagraphStyle(
            'ReportSection',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            spaceBefore=15,
            spaceAfter=10
        )

        # Header
        story.append(Paragraph("Microscopy Analysis Report", title_style))
        story.append(Paragraph(f"<b>Date Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        story.append(Paragraph(f"<b>Operator:</b> {data.get('operator_name', 'System Automated')}", meta_style))
        story.append(Paragraph(f"<b>Total Images Analyzed:</b> {len(data.get('images', []))}", meta_style))
        story.append(Spacer(1, 15))

        # Organism Statistics Summary
        story.append(Paragraph("Detection Summary", section_style))
        summary_data = [["Organism Species", "Count", "Mean Size (µm²)", "Mean Confidence"]]
        for org, stats in data.get("organism_summary", {}).items():
            summary_data.append([
                org,
                str(stats["count"]),
                f"{stats['mean_area']:.2f}",
                f"{stats['mean_confidence']*100:.1f}%"
            ])

        t_summary = Table(summary_data, colWidths=[200, 100, 120, 100])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e7ff')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#1e1b4b')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f9fafb')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ]))
        story.append(t_summary)
        story.append(Spacer(1, 20))

        # Detailed Detections per Image
        story.append(Paragraph("Detailed Image Observations", section_style))
        for img in data.get("images", []):
            story.append(Paragraph(f"<b>Image:</b> {img['name']} ({img['width']}x{img['height']} px)", styles['Heading3']))
            story.append(Paragraph(f"Scale Calibration: {img['scale']:.4f} µm/pixel | Total Annotations: {len(img['detections'])}", meta_style))
            
            det_data = [["Index", "Detected Species", "Confidence", "Area (µm²)", "Perimeter (µm)"]]
            for idx, det in enumerate(img["detections"]):
                det_data.append([
                    str(idx + 1),
                    det["label_class"],
                    f"{det['confidence']*100:.0f}%",
                    f"{det['area_microns']:.2f}",
                    f"{det['perimeter_microns']:.2f}"
                ])
            
            t_det = Table(det_data, colWidths=[50, 200, 80, 100, 90])
            t_det.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f3f4f6')),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#f3f4f6')),
                ('ALIGN', (2,0), (-1,-1), 'CENTER'),
            ]))
            story.append(t_det)
            story.append(Spacer(1, 15))

        doc.build(story)
        return file_path

    def generate_excel_report(self, file_path: str, data: Dict[str, Any]) -> str:
        """
        Generates a structured clinical spreadsheet using openpyxl.
        """
        wb = Workbook()
        
        # Sheet 1: Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary Dashboard"
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="6366F1")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Calibri", size=11)
        header_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Title
        ws_summary["A1"] = "Microscopy Detection Platform - Summary Report"
        ws_summary["A1"].font = title_font
        
        ws_summary["A3"] = "Generated At:"
        ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A4"] = "Total Images Analyzed:"
        ws_summary["B4"] = len(data.get("images", []))
        
        for r in range(3, 5):
            ws_summary[f"A{r}"].font = Font(bold=True)
            
        # Summary Table
        ws_summary["A6"] = "Organism Species"
        ws_summary["B6"] = "Count"
        ws_summary["C6"] = "Mean Area (µm²)"
        ws_summary["D6"] = "Mean Confidence"
        
        for col in ["A", "B", "C", "D"]:
            cell = ws_summary[f"{col}6"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        row_idx = 7
        for org, stats in data.get("organism_summary", {}).items():
            ws_summary.cell(row=row_idx, column=1, value=org).font = normal_font
            ws_summary.cell(row=row_idx, column=2, value=stats["count"]).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row_idx, column=3, value=round(stats["mean_area"], 2))
            ws_summary.cell(row=row_idx, column=4, value=round(stats["mean_confidence"], 4))
            ws_summary.cell(row=row_idx, column=4).number_format = '0.0%'
            
            for c in range(1, 5):
                cell = ws_summary.cell(row=row_idx, column=c)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
            row_idx += 1
            
        # Sheet 2: Detailed Detections
        ws_detail = wb.create_sheet(title="Observations Detail")
        ws_detail["A1"] = "Detailed Microorganism Observations"
        ws_detail["A1"].font = title_font
        
        headers = ["Image Name", "Index", "Class Label", "Confidence", "Area (µm²)", "Perimeter (µm)", "Shape Type"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_detail.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        det_row = 4
        for img in data.get("images", []):
            for idx, det in enumerate(img["detections"]):
                ws_detail.cell(row=det_row, column=1, value=img["name"])
                ws_detail.cell(row=det_row, column=2, value=idx + 1).alignment = Alignment(horizontal="center")
                ws_detail.cell(row=det_row, column=3, value=det["label_class"])
                
                c_cell = ws_detail.cell(row=det_row, column=4, value=det["confidence"])
                c_cell.number_format = '0.0%'
                c_cell.alignment = Alignment(horizontal="center")
                
                ws_detail.cell(row=det_row, column=5, value=det["area_microns"])
                ws_detail.cell(row=det_row, column=6, value=det["perimeter_microns"])
                ws_detail.cell(row=det_row, column=7, value=det["shape_type"])
                
                for c in range(1, 8):
                    cell = ws_detail.cell(row=det_row, column=c)
                    cell.border = thin_border
                    cell.font = normal_font
                    if det_row % 2 == 0:
                        cell.fill = zebra_fill
                det_row += 1
                
        wb.save(file_path)
        return file_path

    def generate_csv_report(self, file_path: str, data: Dict[str, Any]) -> str:
        """
        Generates a raw database dump CSV of all predictions.
        """
        with open(file_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Image Name", "Detection Index", "Organism Label", "Confidence Score", "Area (sq microns)", "Perimeter (microns)", "Shape Type"])
            
            for img in data.get("images", []):
                for idx, det in enumerate(img["detections"]):
                    writer.writerow([
                        img["name"],
                        idx + 1,
                        det["label_class"],
                        det["confidence"],
                        det["area_microns"],
                        det["perimeter_microns"],
                        det["shape_type"]
                    ])
        return file_path

report_service = ReportService()
